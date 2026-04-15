import re
from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_HEADERS = ["DATE", "CJ", "PIECE", "LIBELLE", "DEBIT", "CREDIT"]
EXTRA_HEADERS = [
    "BC",
    "BL",
    "NUM FACTURE",
    "MONTANT FACTURE",
    "ÉCARTS",
    "COMMENTAIRES",
    "CONCLUSION",
    "RÉFÉRENCE",
]


def _to_clean_str(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _normalize_compte(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip().replace(" ", "")
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _normalize_header(value):
    text = _to_clean_str(value).upper()
    text = text.replace("É", "E").replace("È", "E").replace("Ê", "E")
    text = text.replace("À", "A").replace("Ù", "U")
    text = re.sub(r"\s+", " ", text)
    return text


def _num(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        text = str(value).replace(" ", "").replace(",", ".")
        try:
            return float(text)
        except Exception:
            return 0.0


def lire_liste_comptes(fichier):
    df = pd.read_excel(fichier)

    if df.empty:
        return {"comptes": [], "soldes_bg": {}}

    df.columns = [str(c).strip().upper() for c in df.columns]

    compte_col = df.columns[0]

    solde_bg_col = None
    for col in df.columns:
        if col in ["SOLDE BG", "SOLDE_BG", "SOLDEBG"]:
            solde_bg_col = col
            break

    comptes = []
    soldes_bg = {}

    for _, row in df.iterrows():
        compte = _normalize_compte(row.get(compte_col))
        if not compte:
            continue

        comptes.append(compte)

        if solde_bg_col:
            soldes_bg[compte] = _num(row.get(solde_bg_col))
        else:
            soldes_bg[compte] = 0.0

    return {
        "comptes": comptes,
        "soldes_bg": soldes_bg,
    }


def _detect_header_from_row(row_values):
    normalized = [_normalize_header(v) for v in row_values]
    current_map = {}

    for idx, val in enumerate(normalized, start=1):
        if val == "DATE":
            current_map["DATE"] = idx
        elif val == "CJ":
            current_map["CJ"] = idx
        elif val == "PIECE":
            current_map["PIECE"] = idx
        elif val == "LIBELLE":
            current_map["LIBELLE"] = idx
        elif val == "DEBIT":
            current_map["DEBIT"] = idx
        elif val == "CREDIT":
            current_map["CREDIT"] = idx

    score = len(current_map)
    if score >= 4 and ("DATE" in current_map or "LIBELLE" in current_map):
        ordered_existing_headers = [h for h in BASE_HEADERS if h in current_map]
        return current_map, ordered_existing_headers

    return None, None


def _is_account_start_row(row_values, comptes):
    first_non_empty = ""
    for v in row_values:
        s = _to_clean_str(v)
        if s:
            first_non_empty = s
            break

    if not first_non_empty:
        return None

    text = _normalize_compte(first_non_empty)

    match = re.match(r"^(\d+)", text)
    if match:
        first_code = match.group(1)
    else:
        first_code = re.split(r"[\s\-/]+", text)[0]

    if first_code in comptes:
        return first_code

    return None


def _is_total_row(row_values):
    row_text = " ".join(_to_clean_str(v).upper() for v in row_values if _to_clean_str(v))
    return "TOTAL COMPTE" in row_text or row_text.startswith("TOTAL")


def _is_transaction_row(extracted_row):
    debit = extracted_row.get("DEBIT")
    credit = extracted_row.get("CREDIT")

    has_amount = any(x not in (None, "", 0) for x in [debit, credit])
    has_identity = any(
        _to_clean_str(extracted_row.get(k)) != ""
        for k in ["DATE", "PIECE", "LIBELLE", "CJ"]
    )

    return has_identity and has_amount


def extraire_gl_sari_sage(fichier_gl, comptes_input):
    """
    Version optimisée :
    - lecture en read_only
    - parcours une seule fois
    - détection de l'entête pendant le parcours
    """
    if isinstance(comptes_input, dict):
        comptes = comptes_input.get("comptes", [])
        soldes_bg = comptes_input.get("soldes_bg", {})
    else:
        comptes = comptes_input
        soldes_bg = {}

    comptes = [_normalize_compte(c) for c in comptes if _normalize_compte(c)]
    comptes_set = set(comptes)
    soldes_bg = {_normalize_compte(k): v for k, v in soldes_bg.items()}

    resultats = {
        "__meta__": {
            "ordered_existing_headers": [],
            "soldes_bg": soldes_bg,
        }
    }

    wb = load_workbook(fichier_gl, data_only=True, read_only=True)
    ws = wb.active

    header_found = False
    column_map = {}
    ordered_existing_headers = []
    compte_courant = None

    for row_values in ws.iter_rows(values_only=True):
        row_values = list(row_values)

        if not header_found:
            detected_map, detected_headers = _detect_header_from_row(row_values)
            if detected_map:
                column_map = detected_map
                ordered_existing_headers = detected_headers
                resultats["__meta__"]["ordered_existing_headers"] = ordered_existing_headers
                header_found = True
            continue

        started_account = _is_account_start_row(row_values, comptes_set)
        if started_account:
            compte_courant = started_account
            if compte_courant not in resultats:
                resultats[compte_courant] = []
            continue

        if compte_courant is None:
            continue

        if _is_total_row(row_values):
            compte_courant = None
            continue

        extracted = {}
        for header in ordered_existing_headers:
            col_idx = column_map[header] - 1
            extracted[header] = row_values[col_idx] if col_idx < len(row_values) else None

        if _is_transaction_row(extracted):
            resultats[compte_courant].append(extracted)

    wb.close()

    if not header_found:
        raise ValueError("Impossible d'identifier la ligne d'entête du GL.")

    return resultats


def _build_output_headers(existing_headers):
    black_headers = existing_headers.copy()
    if "SOLDE" not in black_headers:
        black_headers.append("SOLDE")

    blue_headers = EXTRA_HEADERS.copy()
    return black_headers + blue_headers, black_headers, blue_headers


def _safe_sheet_title(title):
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    title = _normalize_compte(title)
    for ch in invalid:
        title = title.replace(ch, "_")
    return title[:31] if title else "FEUILLE"


def _copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def _load_model_template(modele, total_cols):
    if not modele:
        return None

    wb_model = load_workbook(modele)
    ws_model = wb_model.active

    header_row_idx = None
    max_scan_rows = min(ws_model.max_row, 80)

    for row_idx in range(1, max_scan_rows + 1):
        row_values = [ws_model.cell(row_idx, col_idx).value for col_idx in range(1, ws_model.max_column + 1)]
        normalized = [_normalize_header(v) for v in row_values]
        found = sum(1 for x in normalized if x in ["DATE", "CJ", "PIECE", "LIBELLE", "DEBIT", "CREDIT"])
        if found >= 3:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return None

    top_end_row = header_row_idx - 1
    if top_end_row < 1:
        return None

    max_col_to_copy = min(max(ws_model.max_column, total_cols), 50)

    merged_ranges = []
    for merged_range in ws_model.merged_cells.ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row

        if max_row <= top_end_row:
            merged_ranges.append((min_row, min_col, max_row, min(max_col, total_cols)))

    methodology_row = None
    for row_idx in range(1, top_end_row + 1):
        for col_idx in range(1, min(ws_model.max_column, total_cols) + 1):
            value = _normalize_header(ws_model.cell(row_idx, col_idx).value)
            if "METHODOLOGIE" in value:
                methodology_row = row_idx
                break
        if methodology_row:
            break

    return {
        "sheet": ws_model,
        "top_end_row": top_end_row,
        "methodology_row": methodology_row,
        "max_col_to_copy": max_col_to_copy,
        "merged_ranges": merged_ranges,
    }


def _apply_model_top_section(ws, compte, template, total_cols):
    src_ws = template["sheet"]
    top_end_row = template["top_end_row"]
    max_col_to_copy = min(template["max_col_to_copy"], total_cols)

    for row_idx in range(1, top_end_row + 1):
        if row_idx in src_ws.row_dimensions:
            ws.row_dimensions[row_idx].height = src_ws.row_dimensions[row_idx].height

        for col_idx in range(1, max_col_to_copy + 1):
            src_cell = src_ws.cell(row=row_idx, column=col_idx)
            dst_cell = ws.cell(row=row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            _copy_cell_style(src_cell, dst_cell)

    for min_row, min_col, max_row, max_col in template["merged_ranges"]:
        if min_col <= total_cols and max_col <= total_cols and min_col <= max_col:
            ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

    meth_row = template.get("methodology_row")
    if meth_row:
        placed = False
        row_ranges = [
            mr for mr in ws.merged_cells.ranges
            if mr.min_row <= meth_row <= mr.max_row
        ]

        if row_ranges:
            max_used_col = max(r.max_col for r in row_ranges)
            candidate_col = max_used_col + 1
            if candidate_col <= total_cols:
                cell = ws.cell(meth_row, candidate_col)
                cell.value = "Aller au MEMO"
                cell.hyperlink = "#MEMO!A1"
                cell.style = "Hyperlink"
                cell.alignment = Alignment(horizontal="left", vertical="center")
                placed = True

        if not placed:
            insertion_test_col = min(total_cols, max(2, total_cols))
            cell = ws.cell(top_end_row + 1, insertion_test_col)
            cell.value = "Aller au MEMO"
            cell.hyperlink = "#MEMO!A1"
            cell.style = "Hyperlink"
            cell.alignment = Alignment(horizontal="left", vertical="center")

    insertion_row = top_end_row + 1
    ws.cell(insertion_row, 1).value = f"Compte : {compte}"
    ws.cell(insertion_row, 1).font = Font(italic=True, bold=True)

    return insertion_row + 1


def _apply_fallback_top_section(ws, compte, total_cols):
    section_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    section_font = Font(bold=True)

    sections = [
        ("OBJECTIFS", "Décrire l'objectif de contrôle."),
        ("METHODOLOGIE", "Décrire la méthodologie."),
        ("RESULTATS", f"Compte : {compte}"),
    ]

    row_idx = 1
    for title, content in sections:
        ws.cell(row_idx, 1).value = title
        ws.cell(row_idx, 1).fill = section_fill
        ws.cell(row_idx, 1).font = section_font
        if total_cols > 1:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        row_idx += 1

        ws.cell(row_idx, 1).value = content
        ws.cell(row_idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        if total_cols > 1:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        row_idx += 2

    ws.cell(4, total_cols).value = "Aller au MEMO"
    ws.cell(4, total_cols).hyperlink = "#MEMO!A1"
    ws.cell(4, total_cols).style = "Hyperlink"

    return row_idx


def _apply_header_style(ws, header_row_idx, black_headers, blue_headers):
    black_fill = PatternFill(fill_type="solid", fgColor="000000")
    blue_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
    white_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(black_headers, start=1):
        cell = ws.cell(header_row_idx, col_idx)
        cell.value = header
        cell.fill = black_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    start_blue = len(black_headers) + 1
    for offset, header in enumerate(blue_headers):
        cell = ws.cell(header_row_idx, start_blue + offset)
        cell.value = header
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border


def _write_account_rows(ws, start_row, rows, existing_headers, full_headers, compte, soldes_bg):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_solde = 0.0
    row_idx = start_row

    for row_data in rows:
        debit = _num(row_data.get("DEBIT"))
        credit = _num(row_data.get("CREDIT"))
        current_solde += debit - credit

        output_row = [row_data.get(header) for header in existing_headers]
        output_row.append(current_solde)
        output_row.extend([""] * len(EXTRA_HEADERS))

        for col_idx, value in enumerate(output_row, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = border

            header_name = full_headers[col_idx - 1]

            if header_name in ["DEBIT", "CREDIT", "SOLDE", "MONTANT FACTURE", "ÉCARTS"]:
                cell.number_format = '#,##0'

            if header_name == "DATE" and isinstance(value, datetime):
                cell.number_format = "dd/mm/yyyy"

            if header_name in ["LIBELLE", "COMMENTAIRES", "CONCLUSION", "RÉFÉRENCE"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_idx += 1

    total_row = row_idx
    ws.cell(total_row, 1).value = "TOTAL"

    if "DEBIT" in full_headers:
        debit_col = full_headers.index("DEBIT") + 1
        total_debit = sum(_num(r.get("DEBIT")) for r in rows)
        ws.cell(total_row, debit_col).value = total_debit
        ws.cell(total_row, debit_col).number_format = '#,##0'

    if "CREDIT" in full_headers:
        credit_col = full_headers.index("CREDIT") + 1
        total_credit = sum(_num(r.get("CREDIT")) for r in rows)
        ws.cell(total_row, credit_col).value = total_credit
        ws.cell(total_row, credit_col).number_format = '#,##0'

    solde_col = full_headers.index("SOLDE") + 1

    solde_gl_abs = abs(current_solde)

    solde_row = total_row + 1
    ws.cell(solde_row, 1).value = "SOLDE GL"
    ws.cell(solde_row, solde_col).value = solde_gl_abs
    ws.cell(solde_row, solde_col).number_format = '#,##0'

    solde_bg = soldes_bg.get(_normalize_compte(compte), 0.0)
    solde_bg_abs = abs(solde_bg)

    solde_bg_row = solde_row + 1
    ws.cell(solde_bg_row, 1).value = "SOLDE BG"
    ws.cell(solde_bg_row, solde_col).value = solde_bg_abs
    ws.cell(solde_bg_row, solde_col).number_format = '#,##0'

    ecart_row = solde_bg_row + 1
    ws.cell(ecart_row, 1).value = "ECART"
    ws.cell(ecart_row, solde_col).value = abs(solde_gl_abs - solde_bg_abs)
    ws.cell(ecart_row, solde_col).number_format = '#,##0'

    for r in [total_row, solde_row, solde_bg_row, ecart_row]:
        for c in range(1, len(full_headers) + 1):
            ws.cell(r, c).border = border
        ws.cell(r, 1).font = Font(bold=True)

    return ecart_row


def _apply_column_widths(ws, full_headers):
    width_by_header = {
        "DATE": 12,
        "CJ": 8,
        "PIECE": 15,
        "LIBELLE": 40,
        "DEBIT": 15,
        "CREDIT": 15,
        "SOLDE": 15,
        "BC": 12,
        "BL": 12,
        "NUM FACTURE": 18,
        "MONTANT FACTURE": 18,
        "ÉCARTS": 15,
        "COMMENTAIRES": 25,
        "CONCLUSION": 20,
        "RÉFÉRENCE": 20,
    }

    for col_idx, header in enumerate(full_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width_by_header.get(header, 15)


def _create_memo_sheet(wb):
    if "MEMO" in wb.sheetnames:
        del wb["MEMO"]

    ws = wb.create_sheet(title="MEMO", index=0)
    ws["A1"] = "MEMO"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Utilisez cette feuille pour centraliser vos notes, justifications, constats et conclusions."
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 100
    return ws


def generer_excel_compte_ex(resultats, chemin_sortie, modele=None):
    meta = resultats.get("__meta__", {})
    existing_headers = meta.get("ordered_existing_headers", [h for h in BASE_HEADERS])
    soldes_bg = meta.get("soldes_bg", {})

    full_headers, black_headers, blue_headers = _build_output_headers(existing_headers)
    total_cols = len(full_headers)

    template = _load_model_template(modele, total_cols) if modele else None

    wb = Workbook()
    wb.remove(wb.active)

    _create_memo_sheet(wb)

    for compte, rows in resultats.items():
        if compte == "__meta__":
            continue

        ws = wb.create_sheet(title=_safe_sheet_title(compte))

        if template:
            header_row_idx = _apply_model_top_section(ws, compte, template, total_cols)
        else:
            header_row_idx = _apply_fallback_top_section(ws, compte, total_cols)

        _apply_header_style(ws, header_row_idx, black_headers, blue_headers)

        if rows:
            _write_account_rows(
                ws=ws,
                start_row=header_row_idx + 1,
                rows=rows,
                existing_headers=existing_headers,
                full_headers=full_headers,
                compte=compte,
                soldes_bg=soldes_bg,
            )

        _apply_column_widths(ws, full_headers)
        ws.freeze_panes = f"A{header_row_idx + 1}"

    wb.save(chemin_sortie)